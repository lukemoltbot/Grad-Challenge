#!/usr/bin/env python3
"""Populate Brave Blossom Capital sheet with scheduling data from vault file 07."""
import openpyxl

# Load workbook preserving formulas
wb = openpyxl.load_workbook('/Users/lukemoltbot/Grad-Challenge/workbooks/Complex_Valuation_Model.xlsx', data_only=False)

# --- Populate Brave Blossom Capital (Sheet 10) ---
ws_cap = wb["Brave Blossom Capital"]

# 1. Enter item names in column A (rows 4-15)
item_names = [
    "Projects / Studies",
    "Exploration Drilling",
    "ROM Bin Upgrade",
    "Ventilation Shafts",
    "Mine Infrastructure",
    "Drift",
    "Drift Conveyor to ROM Stockpile",
    "ROM Stockpile",
    "UG Conveyors Relocation & Reuse",
    "Longwall",
    "Mining Equipment",
    "Infrastructure + CHPP EPCM"
]

for i, name in enumerate(item_names):
    ws_cap.cell(row=4+i, column=1, value=name)

# 2. Enter unit counts per year (D=2027, E=2028, F=2029, G=2030, H=2031, I=2032, J=2033)
# Derived from vault file 07: unit_count = dollar_amount / unit_cost
unit_counts = {
    # row: [D, E, F, G, H, I, J] = [2027, 2028, 2029, 2030, 2031, 2032, 2033]
    4:  [1.0, 1.0, 0.5, 0.5, 0,   0,   0],     # Projects/Studies (5000/unit)
    5:  [1.0, 1.0, 0.5, 0.5, 0,   0,   0],     # Exploration Drilling (5000/unit)
    6:  [0,   0,   0,   0,   1.0, 0.5, 0],     # ROM Bin Upgrade (15000/unit)
    7:  [0,   0,   1.0, 1.0, 0,   0,   0],     # Ventilation Shafts (20000/unit)
    8:  [0,   0,   1.0, 1.0, 0,   0,   0],     # Mine Infrastructure (5000/unit)
    9:  [0,   0,   1.0, 1.0, 0,   0,   0],     # Drift (12000/unit)
    10: [0,   0,   1.0, 1.0, 0,   0,   0],     # Drift Conveyor (17000/unit)
    11: [0,   0,   0,   1.0, 1.0, 0,   0],     # ROM Stockpile (5000/unit)
    12: [0,   0,   0,   1.0, 1.0, 0,   0],     # UG Conveyors (20000/unit)
    13: [0,   0,   0,   0,   0,   1.0, 0],     # Longwall (190000/unit)
    14: [0,   0,   0,   0,   1.0, 0,   0],     # Mining Equipment (69900/unit)
    15: [0,   0,   0,   0,   0,   1.0, 1.0],   # CHPP EPCM (24000/unit)
}

for row, counts in unit_counts.items():
    for col_idx, count in enumerate(counts):
        col = 4 + col_idx  # D=4, E=5, ..., J=10
        ws_cap.cell(row=row, column=col, value=count)

# Save populated version
output_path = '/Users/lukemoltbot/Grad-Challenge/workbooks/Complex_Valuation_Model_POPULATED.xlsx'
wb.save(output_path)
print(f"Saved populated workbook to: {output_path}")

# --- Verify by reading back ---
wb2 = openpyxl.load_workbook(output_path, data_only=False)
ws2 = wb2["Brave Blossom Capital"]

print("\n=== Verification: Item Names ===")
for row in range(4, 16):
    name = ws2.cell(row=row, column=1).value
    cost = ws2.cell(row=row, column=3).value
    print(f"  Row {row}: {name} | Unit cost: {cost}")

print("\n=== Verification: Unit Counts (D-J = 2027-2033) ===")
for row in range(4, 16):
    counts = [ws2.cell(row=row, column=col).value for col in range(4, 11)]
    print(f"  Row {row} ({ws2.cell(row=row, column=1).value}): {counts}")

# --- Compute expected capital schedule ---
print("\n=== Computed Capital Schedule (AUD$k, before contingency) ===")
unit_costs_map = {4:5000, 5:5000, 6:15000, 7:20000, 8:5000, 9:12000,
             10:17000, 11:5000, 12:20000, 13:190000, 14:69900, 15:24000}

years = [2027, 2028, 2029, 2030, 2031, 2032, 2033]
year_totals = [0] * 7

for row, counts in unit_counts.items():
    cost = unit_costs_map[row]
    for i, count in enumerate(counts):
        year_totals[i] += cost * count

print(f"{'Year':<8} {'Total (AUD$k)':<15} {'Contingency 30%':<18} {'Total Capital':<15}")
for i, year in enumerate(years):
    cont = year_totals[i] * 0.30
    total = year_totals[i] + cont
    print(f"{year:<8} {year_totals[i]:<15,.0f} {cont:<18,.0f} {total:<15,.0f}")

grand_total = sum(year_totals)
grand_cont = grand_total * 0.30
grand_capital = grand_total + grand_cont
print(f"\n{'TOTAL':<8} {grand_total:<15,.0f} {grand_cont:<18,.0f} {grand_capital:<15,.0f}")
print(f"\nWithout contingency: ${grand_total/1000:.1f}M")
print(f"With 30% contingency: ${grand_capital/1000:.1f}M")
print(f"Vault estimate (excl contingency): $388.9M")
print(f"Vault estimate (incl 30% contingency): ${388.9*1.3:.1f}M")
