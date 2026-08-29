"""Comprehensive audit of Complex_Valuation_Model.xlsx.
Dumps raw, structured data for the audit report.
"""
import json
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

WB_PATH = "/Users/lukemoltbot/Grad-Challenge/workbooks/Complex_Valuation_Model.xlsx"

# data_only=False keeps formulas; we also load data_only=True for values
wb_f = load_workbook(WB_PATH, data_only=False)
wb_v = load_workbook(WB_PATH, data_only=True)

print("=" * 80)
print("SHEET COUNT:", len(wb_f.sheetnames))
print("=" * 80)
print("SHEET NAMES (repr to expose trailing spaces):")
for i, name in enumerate(wb_f.sheetnames, 1):
    print(f"  [{i:>2}] {name!r}  (len={len(name)})")

# Collect per-sheet stats
sheet_stats = []
cross_sheet_links = []  # list of (sheet, cell, formula, referenced sheets)

# Pattern to find cross-sheet refs like 'Sheet Name'!A1 or SheetName!A1
# Note: sheet names can contain spaces and need quotes in Excel
ref_pattern = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_\-.]+))!([A-Za-z0-9_$:.{}]+)")

for idx, name in enumerate(wb_f.sheetnames, 1):
    ws_f = wb_f[name]
    ws_v = wb_v[name]
    dims = ws_f.dimensions  # e.g. 'A1:Z100'
    max_row = ws_f.max_row
    max_col = ws_f.max_column
    non_empty = 0
    formula_cells = 0
    formula_samples = []
    numeric_cells = 0
    text_cells = 0
    blank_cells = 0
    # iterate all cells in used range
    for row in ws_f.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            non_empty += 1
            if isinstance(v, str) and v.startswith("="):
                formula_cells += 1
                if len(formula_samples) < 5:
                    formula_samples.append(f"{cell.coordinate}: {v}")
                # check cross-sheet
                for m in ref_pattern.finditer(v):
                    ref_sheet = m.group(1) or m.group(2)
                    # exclude self-references to the same sheet
                    if ref_sheet != name:
                        cross_sheet_links.append((name, cell.coordinate, v, ref_sheet))
            elif isinstance(v, (int, float)):
                numeric_cells += 1
            else:
                text_cells += 1
    sheet_stats.append({
        "index": idx,
        "name": name,
        "name_repr": repr(name),
        "dimensions": dims,
        "max_row": max_row,
        "max_col": max_col,
        "non_empty": non_empty,
        "formula_cells": formula_cells,
        "numeric_cells": numeric_cells,
        "text_cells": text_cells,
        "blank_in_range": (max_row * max_col) - non_empty,
        "formula_samples": formula_samples,
    })

print()
print("=" * 80)
print("PER-SHEET STATS")
print("=" * 80)
print(f"{'#':>3} {'Name':<40} {'Dims':<14} {'NonEmpty':>9} {'Formulas':>9}")
for s in sheet_stats:
    print(f"{s['index']:>3} {s['name'][:40]:<40} {s['dimensions']:<14} {s['non_empty']:>9} {s['formula_cells']:>9}")

# Detailed per-sheet dump
print()
print("=" * 80)
print("DETAILED PER-SHEET")
print("=" * 80)
for s in sheet_stats:
    print(f"\n--- [{s['index']}] {s['name']!r} ---")
    print(f"  Dimensions: {s['dimensions']}  (rows={s['max_row']}, cols={s['max_col']})")
    print(f"  Non-empty: {s['non_empty']}  Numeric: {s['numeric_cells']}  Text: {s['text_cells']}  Formulas: {s['formula_cells']}")
    print(f"  Blank cells in used range: {s['blank_in_range']}")
    if s['formula_samples']:
        print("  Formula samples:")
        for fs in s['formula_samples']:
            print(f"    {fs}")
    else:
        print("  (no formulas)")

# Cross-sheet links
print()
print("=" * 80)
print("CROSS-SHEET LINKS")
print("=" * 80)
if not cross_sheet_links:
    print("  NONE FOUND")
else:
    # dedupe by (sheet, ref_sheet) for summary
    by_source = {}
    for src, coord, formula, ref in cross_sheet_links:
        by_source.setdefault(src, set()).add(ref)
    print("  Summary (source sheet -> referenced sheets):")
    for src, refs in by_source.items():
        print(f"    {src!r} -> {sorted(refs)}")
    print()
    print("  Full link list (first 60):")
    for src, coord, formula, ref in cross_sheet_links[:60]:
        fshort = formula if len(formula) <= 90 else formula[:87] + "..."
        print(f"    [{src}] {coord} refs {ref!r}: {fshort}")
    if len(cross_sheet_links) > 60:
        print(f"    ... and {len(cross_sheet_links)-60} more")

# Detect BROKEN cross-sheet links (referenced sheet doesn't exist)
print()
print("=" * 80)
print("BROKEN CROSS-SHEET LINKS (ref sheet not in workbook)")
print("=" * 80)
sheet_set = set(wb_f.sheetnames)
broken = []
for src, coord, formula, ref in cross_sheet_links:
    # handle potential ref variants (some sheets may have trailing spaces)
    if ref not in sheet_set:
        # try stripped/with-space variants
        candidates = [ref, ref.strip(), ref + " "]
        if not any(c in sheet_set for c in candidates):
            broken.append((src, coord, ref, formula))
if not broken:
    print("  NONE - all referenced sheets exist")
else:
    for src, coord, ref, formula in broken:
        print(f"    [{src}] {coord} -> {ref!r}  formula: {formula[:80]}")

print()
print("AUDIT SCRIPT COMPLETE")
