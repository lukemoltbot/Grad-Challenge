"""Deep dive on OC Clean Capital, Springbok, Brave Blossom, Decommissioned Mine."""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

WB_PATH = "/Users/lukemoltbot/Grad-Challenge/workbooks/Complex_Valuation_Model.xlsx"
wb_f = load_workbook(WB_PATH, data_only=False)
wb_v = load_workbook(WB_PATH, data_only=True)

def dump_rows(name, max_row=None, max_cols=15, show_formulas=True):
    ws_f = wb_f[name]
    ws_v = wb_v[name]
    mr = max_row or ws_f.max_row
    mc = min(max_cols or ws_f.max_column, ws_f.max_column)
    print(f"  Dimensions: {ws_f.dimensions}  rows={ws_f.max_row} cols={ws_f.max_column} (showing up to {mc} cols, {mr} rows)")
    for r in range(1, mr+1):
        cells = []
        for c in range(1, mc+1):
            cf = ws_f.cell(row=r, column=c)
            cv = ws_v.cell(row=r, column=c)
            if cf.value is None and cv.value is None:
                continue
            cl = get_column_letter(c)
            fval = cf.value
            vval = cv.value
            if isinstance(fval, str) and fval.startswith("="):
                if show_formulas:
                    cells.append(f"{cl}{r}={fval}")
                else:
                    cells.append(f"{cl}{r}=[{vval}]")
            else:
                cells.append(f"{cl}{r}={fval!r}")
        if cells:
            print(f"  R{r:>3}: " + " | ".join(cells[:15]))

# ============================================================
# OC Clean Capital (sheet 13) and OC Clean (sheet 12)
# ============================================================
print("=" * 90)
print("SHEET 13: 'OC Clean Capital'")
print("=" * 90)
dump_rows("OC Clean Capital", max_cols=20)
print()
print("=" * 90)
print("SHEET 12: 'OC Clean' — first 30 rows + rows 140-153 (Safeguard area)")
print("=" * 90)
dump_rows("OC Clean", max_row=30, max_cols=12)
print("  ... (rows 140-153, the Safeguard/baseline area) ...")
ws = wb_f["OC Clean"]
for r in range(140, min(ws.max_row+1, 154)):
    cells = []
    for c in range(1, min(13, ws.max_column+1)):
        cf = ws.cell(row=r, column=c)
        cv = wb_v["OC Clean"].cell(row=r, column=c)
        if cf.value is None and cv.value is None:
            continue
        cl = get_column_letter(c)
        fval = cf.value
        vval = cv.value
        if isinstance(fval, str) and fval.startswith("="):
            cells.append(f"{cl}{r}={fval}")
        else:
            cells.append(f"{cl}{r}={fval!r}")
    if cells:
        print(f"  R{r:>3}: " + " | ".join(cells[:12]))

# ============================================================
# Springbok (sheet 7) — full label column + Safeguard area (rows 140-161)
# ============================================================
print()
print("=" * 90)
print("SHEET 7: 'Springbok' — label column (A:C) full + rows 140-161 (Safeguard)")
print("=" * 90)
ws = wb_f["Springbok"]
ws_v7 = wb_v["Springbok"]
print(f"  Dimensions: {ws.dimensions}  rows={ws.max_row} cols={ws.max_column}")
print("  Label rows (col A-B-C) with values:")
for r in range(1, ws.max_row+1):
    cells = []
    for c in range(1, 4):  # A, B, C
        cf = ws.cell(row=r, column=c)
        cv = ws_v7.cell(row=r, column=c)
        if cf.value is None and cv.value is None:
            continue
        cl = get_column_letter(c)
        fval = cf.value
        vval = cv.value
        if isinstance(fval, str) and fval.startswith("="):
            cells.append(f"{cl}{r}={fval}")
        else:
            cells.append(f"{cl}{r}={fval!r}")
    if cells:
        print(f"  R{r:>3}: " + " | ".join(cells))
print()
print("  --- Rows 140-161 (Safeguard Mechanism area), cols A-M ---")
for r in range(140, min(ws.max_row+1, 162)):
    cells = []
    for c in range(1, 14):
        cf = ws.cell(row=r, column=c)
        cv = ws_v7.cell(row=r, column=c)
        if cf.value is None and cv.value is None:
            continue
        cl = get_column_letter(c)
        fval = cf.value
        vval = cv.value
        if isinstance(fval, str) and fval.startswith("="):
            cells.append(f"{cl}{r}={fval}=[{vval}]")
        else:
            cells.append(f"{cl}{r}={fval!r}")
    if cells:
        print(f"  R{r:>3}: " + " | ".join(cells[:14]))

# ============================================================
# Brave Blossom (sheet 9) — label column + Safeguard area (rows 140-150)
# ============================================================
print()
print("=" * 90)
print("SHEET 9: 'Brave Blossom ' (trailing space) — label column + rows 140-150")
print("=" * 90)
ws = wb_f["Brave Blossom "]
ws_v9 = wb_v["Brave Blossom "]
print(f"  Dimensions: {ws.dimensions}  rows={ws.max_row} cols={ws.max_column}")
print("  Label rows (col A-B-C):")
for r in range(1, ws.max_row+1):
    cells = []
    for c in range(1, 4):
        cf = ws.cell(row=r, column=c)
        cv = ws_v9.cell(row=r, column=c)
        if cf.value is None and cv.value is None:
            continue
        cl = get_column_letter(c)
        fval = cf.value
        vval = cv.value
        if isinstance(fval, str) and fval.startswith("="):
            cells.append(f"{cl}{r}={fval}")
        else:
            cells.append(f"{cl}{r}={fval!r}")
    if cells:
        print(f"  R{r:>3}: " + " | ".join(cells))
print()
print("  --- Rows 140-150 (Safeguard area), cols A-M ---")
for r in range(140, min(ws.max_row+1, 151)):
    cells = []
    for c in range(1, 14):
        cf = ws.cell(row=r, column=c)
        cv = ws_v9.cell(row=r, column=c)
        if cf.value is None and cv.value is None:
            continue
        cl = get_column_letter(c)
        fval = cf.value
        vval = cv.value
        if isinstance(fval, str) and fval.startswith("="):
            cells.append(f"{cl}{r}={fval}=[{vval}]")
        else:
            cells.append(f"{cl}{r}={fval!r}")
    if cells:
        print(f"  R{r:>3}: " + " | ".join(cells[:14]))

# ============================================================
# Decommissioned Mine (sheet 15)
# ============================================================
print()
print("=" * 90)
print("SHEET 15: 'Decommissioned Mine' — full structural dump (cols A-J, all rows)")
print("=" * 90)
ws = wb_f["Decommissioned Mine"]
ws_v15 = wb_v["Decommissioned Mine"]
print(f"  Dimensions: {ws.dimensions}  rows={ws.max_row} cols={ws.max_column}")
for r in range(1, ws.max_row+1):
    cells = []
    for c in range(1, min(11, ws.max_column+1)):  # A-J
        cf = ws.cell(row=r, column=c)
        cv = ws_v15.cell(row=r, column=c)
        if cf.value is None and cv.value is None:
            continue
        cl = get_column_letter(c)
        fval = cf.value
        vval = cv.value
        if isinstance(fval, str) and fval.startswith("="):
            cells.append(f"{cl}{r}={fval}=[{vval}]")
        else:
            cells.append(f"{cl}{r}={fval!r}")
    if cells:
        print(f"  R{r:>3}: " + " | ".join(cells[:10]))

print()
print("DEEP DIVE COMPLETE")
