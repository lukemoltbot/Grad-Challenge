"""Final targeted data: Safeguard baselines, NPV cells, Summary, Assumptions."""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

WB_PATH = "/Users/lukemoltbot/Grad-Challenge/workbooks/Complex_Valuation_Model.xlsx"
wb_f = load_workbook(WB_PATH, data_only=False)
wb_v = load_workbook(WB_PATH, data_only=True)

def row_dump(name, r, max_col=15, show_both=True):
    ws_f = wb_f[name]
    ws_v = wb_v[name]
    cells = []
    for c in range(1, min(max_col+1, ws_f.max_column+1)):
        cf = ws_f.cell(row=r, column=c).value
        cv = ws_v.cell(row=r, column=c).value
        if cf is None and cv is None:
            continue
        cl = get_column_letter(c)
        if isinstance(cf, str) and cf.startswith("="):
            if show_both:
                cells.append(f"{cl}{r}={cf} → [{cv}]")
            else:
                cells.append(f"{cl}{r}=[{cv}]")
        else:
            cells.append(f"{cl}{r}={cf!r}")
    return cells

# Springbok Safeguard / Carbon area rows 21-29
print("=" * 80)
print("SPRINGBOK (sheet 7) — Carbon/Safeguard rows 21-29")
print("=" * 80)
for r in range(21, 30):
    cells = row_dump("Springbok", r, max_col=12)
    if cells:
        print(f"  R{r}: " + " | ".join(cells[:10]))

# Brave Blossom Safeguard / Carbon area rows 22-29
print()
print("=" * 80)
print("BRAVE BLOSSOM (sheet 9) — Carbon/Safeguard rows 22-29")
print("=" * 80)
for r in range(22, 30):
    cells = row_dump("Brave Blossom ", r, max_col=12)
    if cells:
        print(f"  R{r}: " + " | ".join(cells[:10]))

# OC Clean Safeguard / Carbon area rows 26-30
print()
print("=" * 80)
print("OC CLEAN (sheet 12) — Carbon/Safeguard rows 26-30")
print("=" * 80)
for r in range(26, 31):
    cells = row_dump("OC Clean", r, max_col=12)
    if cells:
        print(f"  R{r}: " + " | ".join(cells[:10]))

# NPV and key result cells
print()
print("=" * 80)
print("KEY RESULT CELLS (NPV, IRR, Payback) — cached values")
print("=" * 80)
for sheet, label in [("Springbok", "Springbok"), ("Brave Blossom ", "Brave Blossom")]:
    ws_v_s = wb_v[sheet]
    ws_f_s = wb_f[sheet]
    for r in [126, 127, 128]:
        a = ws_f_s.cell(row=r, column=1).value
        c_val = ws_v_s.cell(row=r, column=3).value
        c_form = ws_f_s.cell(row=r, column=3).value
        print(f"  [{label}] R{r} A={a!r}  C formula={c_form!r}  C value={c_val!r}")

# Summary sheet
print()
print("=" * 80)
print("SUMMARY (sheet 4) — structural dump (cols A-I, rows 1-57)")
print("=" * 80)
ws = wb_f["Summary"]
ws_v4 = wb_v["Summary"]
print(f"  Dimensions: {ws.dimensions}")
for r in range(1, ws.max_row + 1):
    cells = []
    for c in range(1, min(10, ws.max_column + 1)):
        cf = ws.cell(row=r, column=c).value
        cv = ws_v4.cell(row=r, column=c).value
        if cf is None and cv is None:
            continue
        cl = get_column_letter(c)
        if isinstance(cf, str) and cf.startswith("="):
            cells.append(f"{cl}{r}={cf} → [{cv}]")
        else:
            cells.append(f"{cl}{r}={cf!r}")
    if cells:
        print(f"  R{r:>3}: " + " | ".join(cells[:9]))

# Assumptions key rows
print()
print("=" * 80)
print("ASSUMPTIONS (sheet 1) — key rows (cols A-C)")
print("=" * 80)
ws = wb_f["Assumptions"]
ws_v1 = wb_v["Assumptions"]
for r in range(1, ws.max_row + 1):
    cells = []
    for c in range(1, min(5, ws.max_column + 1)):
        cf = ws.cell(row=r, column=c).value
        cv = ws_v1.cell(row=r, column=c).value
        if cf is None and cv is None:
            continue
        cl = get_column_letter(c)
        if isinstance(cf, str) and cf.startswith("="):
            cells.append(f"{cl}{r}={cf}")
        else:
            cells.append(f"{cl}{r}={cf!r}")
    if cells:
        print(f"  R{r:>3}: " + " | ".join(cells[:4]))

# Carbon sheet
print()
print("=" * 80)
print("CARBON (sheet 14) — full dump (all rows, cols A-J)")
print("=" * 80)
ws = wb_f["Carbon"]
ws_v14 = wb_v["Carbon"]
for r in range(1, ws.max_row + 1):
    cells = []
    for c in range(1, min(11, ws.max_column + 1)):
        cf = ws.cell(row=r, column=c).value
        cv = ws_v14.cell(row=r, column=c).value
        if cf is None and cv is None:
            continue
        cl = get_column_letter(c)
        if isinstance(cf, str) and cf.startswith("="):
            cells.append(f"{cl}{r}={cf} → [{cv}]")
        else:
            cells.append(f"{cl}{r}={cf!r}")
    if cells:
        print(f"  R{r}: " + " | ".join(cells[:10]))

# OC Clean Capital value distribution
print()
print("=" * 80)
print("OC CLEAN CAPITAL (sheet 13) — numeric value distribution")
print("=" * 80)
ws_f13 = wb_f["OC Clean Capital"]
ws_v13 = wb_v["OC Clean Capital"]
zeros = nonzeros = blanks = 0
for r in range(1, ws_v13.max_row + 1):
    for c in range(1, ws_v13.max_column + 1):
        cf = ws_f13.cell(row=r, column=c)
        cv = ws_v13.cell(row=r, column=c)
        if cf.value is None:
            blanks += 1
            continue
        v = cv.value
        if isinstance(v, (int, float)):
            if v == 0:
                zeros += 1
            else:
                nonzeros += 1
        elif v is None:
            zeros += 1
print(f"  Zeros: {zeros}  Non-zeros: {nonzeros}  Blanks: {blanks}")

# List sheet
print()
print("=" * 80)
print("LIST (sheet 3) — full dump")
print("=" * 80)
ws = wb_f["List"]
for r in range(1, ws.max_row + 1):
    cells = []
    for c in range(1, ws.max_column +1):
        v = ws.cell(row=r, column=c).value
        if v is None:
            continue
        cl = get_column_letter(c)
        cells.append(f"{cl}{r}={v!r}")
    if cells:
        print(f"  R{r}: " + " | ".join(cells))

# Springbok Assumptions sheet
print()
print("=" * 80)
print("SPRINGBOK ASSUMPTIONS (sheet 6) — full dump")
print("=" * 80)
ws = wb_f["Springbok Assumptions"]
ws_v6 = wb_v["Springbok Assumptions"]
for r in range(1, ws.max_row + 1):
    cells = []
    for c in range(1, ws.max_column+1):
        cf = ws.cell(row=r, column=c).value
        cv = ws_v6.cell(row=r, column=c).value
        if cf is None and cv is None:
            continue
        cl = get_column_letter(c)
        cells.append(f"{cl}{r}={cf!r}")
    if cells:
        print(f"  R{r}: " + " | ".join(cells))

print()
print("TARGETED DATA COMPLETE")
