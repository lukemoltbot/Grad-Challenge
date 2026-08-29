"""Deep structural analysis of key sheets for the audit report."""
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

WB_PATH = "/Users/lukemoltbot/Grad-Challenge/workbooks/Complex_Valuation_Model.xlsx"
wb_f = load_workbook(WB_PATH, data_only=False)
wb_v = load_workbook(WB_PATH, data_only=True)

def dump_sheet_rows(name, max_row=None, max_col=None, value_only=True, ws_v=None):
    """Dump rows with cell coords and values."""
    ws = ws_v if ws_v else wb_f[name]
    mr = max_row or ws.max_row
    mc = max_col or ws.max_column
    rows_out = []
    for r in range(1, mr+1):
        row_vals = []
        for c in range(1, mc+1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                continue
            col_letter = get_column_letter(c)
            row_vals.append((cell.coordinate, v))
        if row_vals:
            rows_out.append((r, row_vals))
    return rows_out

def safe_get(ws, coord):
    try:
        return ws[coord].value
    except Exception:
        return None

# ============================================================
# 1. Brave Blossom Capital (sheet 10) — structural deep-dive
# ============================================================
print("=" * 90)
print("SHEET 10: 'Brave Blossom Capital' — STRUCTURAL DUMP (formulas + values)")
print("=" * 90)
ws_f_bb = wb_f["Brave Blossom Capital"]
ws_v_bb = wb_v["Brave Blossom Capital"]
print(f"Dimensions: {ws_f_bb.dimensions}  rows={ws_f_bb.max_row}  cols={ws_f_bb.max_column}")
print()
# Dump row by row: col A label + first few year cells with formula AND cached value
for r in range(1, ws_f_bb.max_row + 1):
    row_cells_f = []
    row_cells_v = []
    for c in range(1, ws_f_bb.max_column + 1):
        cf = ws_f_bb.cell(row=r, column=c)
        cv = ws_v_bb.cell(row=r, column=c)
        if cf.value is None and cv.value is None:
            continue
        cl = get_column_letter(c)
        fval = cf.value
        vval = cv.value
        if isinstance(fval, str) and fval.startswith("="):
            row_cells_f.append(f"{cl}{r}={fval}")
            row_cells_v.append(f"{cl}{r}=[{vval}]")
        else:
            row_cells_f.append(f"{cl}{r}={fval!r}")
            row_cells_v.append(f"{cl}{r}={vval!r}")
    if row_cells_f:
        print(f"  R{r:>2}: " + " | ".join(row_cells_f[:12]))
        if row_cells_v != row_cells_f:
            print(f"        VALUES: " + " | ".join(row_cells_v[:12]))

# Count zero vs populated numeric in value workbook
print()
print("Brave Blossom Capital — numeric value distribution (from cached values):")
zeros = 0
nonzeros = 0
blanks = 0
formulas_resolved_to_zero = 0
formulas_resolved_nonzero = 0
for r in range(1, ws_v_bb.max_row + 1):
    for c in range(1, ws_v_bb.max_column + 1):
        cf = ws_f_bb.cell(row=r, column=c)
        cv = ws_v_bb.cell(row=r, column=c)
        if cf.value is None:
            blanks += 1
            continue
        is_formula = isinstance(cf.value, str) and cf.value.startswith("=")
        v = cv.value
        if isinstance(v, (int, float)):
            if v == 0:
                zeros += 1
                if is_formula:
                    formulas_resolved_to_zero += 1
            else:
                nonzeros += 1
                if is_formula:
                    formulas_resolved_nonzero += 1
        else:
            # text or None (formula unresolved)
            if v is None and is_formula:
                formulas_resolved_to_zero += 1  # treat unresolved as zero-ish
print(f"  Zeros (numeric==0): {zeros}  (of which formula-derived: {formulas_resolved_to_zero})")
print(f"  Non-zero numerics: {nonzeros}  (of which formula-derived: {formulas_resolved_nonzero})")
print(f"  Blanks (None in both): {blanks}")

# ============================================================
# 2. SpringbokCapital (sheet 8) — template comparison
# ============================================================
print()
print("=" * 90)
print("SHEET 8: 'SpringbokCapital' — STRUCTURAL DUMP (formulas + values)")
print("=" * 90)
ws_f_sb = wb_f["SpringbokCapital"]
ws_v_sb = wb_v["SpringbokCapital"]
print(f"Dimensions: {ws_f_sb.dimensions}  rows={ws_f_sb.max_row}  cols={ws_f_sb.max_column}")
print()
for r in range(1, ws_f_sb.max_row + 1):
    row_cells_f = []
    row_cells_v = []
    for c in range(1, ws_f_sb.max_column + 1):
        cf = ws_f_sb.cell(row=r, column=c)
        cv = ws_v_sb.cell(row=r, column=c)
        if cf.value is None and cv.value is None:
            continue
        cl = get_column_letter(c)
        fval = cf.value
        vval = cv.value
        if isinstance(fval, str) and fval.startswith("="):
            row_cells_f.append(f"{cl}{r}={fval}")
            row_cells_v.append(f"{cl}{r}=[{vval}]")
        else:
            row_cells_f.append(f"{cl}{r}={fval!r}")
            row_cells_v.append(f"{cl}{r}={vval!r}")
    if row_cells_f:
        print(f"  R{r:>2}: " + " | ".join(row_cells_f[:14]))
        if row_cells_v != row_cells_f:
            print(f"        VALUES: " + " | ".join(row_cells_v[:14]))

# value distribution
print()
print("SpringbokCapital — numeric value distribution:")
zeros = nonzeros = blanks = 0
f_zero = f_nonzero = 0
for r in range(1, ws_v_sb.max_row + 1):
    for c in range(1, ws_v_sb.max_column + 1):
        cf = ws_f_sb.cell(row=r, column=c)
        cv = ws_v_sb.cell(row=r, column=c)
        if cf.value is None:
            blanks += 1
            continue
        is_f = isinstance(cf.value, str) and cf.value.startswith("=")
        v = cv.value
        if isinstance(v, (int, float)):
            if v == 0:
                zeros += 1
                if is_f: f_zero += 1
            else:
                nonzeros += 1
                if is_f: f_nonzero += 1
        else:
            if v is None and is_f:
                f_zero += 1
print(f"  Zeros: {zeros} (formula-derived: {f_zero})  Non-zeros: {nonzeros} (formula-derived: {f_nonzero})  Blanks: {blanks}")

print()
print("STRUCTURAL DUMPS COMPLETE")
